# Izail Chamberlain
# UWYO COSC 1010
# Submission Date: 11/19/24
# Homework 5
# Lab Section: 11
# Sources, people worked with, help given to: Used AI generated pixel art as reference to avoid copyright issues
# N/A
"""Homework 5
UWYO COSC 1010

1 Homework 5 - Pixel Art
1.1 Overview
For this program, you will be creating a program to generate pixel art of your choice in a spreadsheet (Excel). Using the openpyxl library outlined in the lecture, you need to create a program that will generate a workbook, set the corresponding cells to be squares (or as close as you can get), and fill in the squares as needed. There is a reference image below. While there is distortion in the image, the cells are (and should be in your program) squares. You may not submit the example pixel art for your project.

1.2 Notes
Your code MUST contain all it needs to generate the image.
This means you cannot include your practice Excel file and copy it over to a new file to complete the assignment.
If you use the scraper outlined below, you can only use it to generate data structures for use in your code. The scrapper CANNOT be used in your program submission
You will likely need to experiment with how you set the width/height until you find squares.
Recall the height is defined by points where 1 point is equivalent to 1/72 of an inch and can be a value of 0-409
Width is can be set to a number 0-255 and denotes how many characters at 11 point font would be shown
The point for the font is the same as for the height
Roughly you can use a formula of row height = 6*column width
A good approach is to create a spreadsheet (or use grid paper) first to figure out what your pixel art will look like.
Once you figure out the layout of your image, you should know what cells are filled and how they are filled.
That information needs to be translated into code.
If you aren’t artistic, you can find pixel art to replicate.
You will need to submit the original image with your solution and cite it.
They picture doesn't need to be recognizable as something if you don't want
Your art must have at least 5 colors and be made of at least 100 cells.
Some students have had issues with local installs of Excel not displaying the images, you can try uploading it to your online Excel instance, or use an alternative like LibreOffice
Example image: [Link from Zelda franchise]
link.png
1.3 Scraper
How to get the coordinates you need to color into your program is up to you. Once you’ve figured out what your image is and the layout, you can manually enter all the cells to be colored into your program to be held in a data structure of some kind.

NOTE: This part is optional; it is harder programming work but will save you tedium. Alternatively, if you made the image in Excel first to work out the layout, you can scrape that file to generate the needed information for your program to make the file. Above, it is stated that your generation program must be standalone. You can use a scraper to get a list of all coordinates and their respective colors, print them out, and store that information in some data structure (like lists or a dictionary) in your generation program. That is to say, you can use a program to generate the data structures that you will then put in your image creation program.

You can determine if a cell is filled with if cell.fill.patternType != None.
You can pull the cell’s color with cell.fill.fgColor.rgb.
The cell’s coordinate can be found with cell.coordinate.
To scrape, remember you can select the cells based on a rectangular portion or via a slice of rows or columns.
Remember to cast it to a tuple to work with it most effectively.
You can then iterate through your selected cells to pull their fill color and coordinate.
wb = openpyxl.load_workbook("your file name") # open the workbook to scrape
sheet = wb.active # get the active sheet from it
cells = tuple(sheet[’A1’:lowerbound]) # get your collection of cells in their rows
color_dict = {} # create a dictionary to hold the colors and coordinates
for row in cells: # go through all the rows
    for cell in row:
        check if the cell is filled
        pull the cell color out of cell, store it
        pull the coordinate out of cell, store it
        see if the color already in the dictionary
        if yes, add the coordinate to a list that corresponds with the color as a key
        (the list would be the value, holding all coordinates that use that color)
        else
        add the color as a key to the dictionary, give it a value of a list with only
        the first coord found
print out your dictionary in a format you can use in your other programs
1.4 Submission
Name your file YourLastName_HW5.py

You MUST include your name and lab section at the top of your file, as well as any sources used."""




from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Initializes workbook and get the active sheet
wb = Workbook()
ws = wb.active

# Defines pixel art as dictionary with colors and coordinates
# Hex codes/colors
pixel_art = {
    "000000": [  # Black (outline)
        "F6", "G6", "H6", "I6", "J6", "K6", "L6", "M6", 
        "E7", "N7", 
        "D8", "O8", 
        "C9", "P9", 
        "B10", "Q10", 
        "C11", "P11", 
        "D12", "O12", 
        "E13", "N13", 
        "F14", "G14", "H14", "I14", "J14", "K14", "L14", "M14",
    ],
    "FF00FF": [  # Magenta (mushroom cap)
        "F7", "G7", "H7", "I7", "J7", "K7", "L7", "M7",
        "F8", "G8", "H8", "I8", "J8", "K8", "L8", "M8",
        "E9", "F9", "G9", "H9", "I9", "J9", "K9", "L9", "M9", "N9",
        "E10", "F10", "G10", "H10", "I10", "J10", "K10", "L10", "M10", "N10",
        "D11", "E11", "F11", "G11", "H11", "I11", "J11", "K11", "L11", "M11", "N11", "O11",
        "E12", "F12", "G12", "H12", "I12", "J12", "K12", "L12", "M12", "N12",
        "F13", "G13", "H13", "I13", "J13", "K13", "L13", "M13",
    ],
    "FFFFFF": [  # White (dots on the mushroom cap)
        "H8", "I8",
        "K9", "L9",
        "G10", "H10",
    ],
    "D2B48C": [  # Tan (mushroom stalk)
        "I15", "J15", 
        "I16", "J16", 
        "I17", "J17", 
        "H18", "I18", "J18", "K18",
    ],
    "FFFFFF": [  # Background (white)
        # Fills background with blank cells
    ],
}

# Sets row height/column width for square cells
for row in range(1, 20):  # Adjusts based on image size
    ws.row_dimensions[row].height = 15  # Adjusted for square appearance
for col in range(1, 20):  # Adjusts based on image size
    ws.column_dimensions[chr(64 + col)].width = 2.5  # Adjusted for square appearance

# Applies the colors to the corresponding cells
for color, cells in pixel_art.items():
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in cells:
        ws[cell].fill = fill

# Saves the workbook
wb.save("Pixel_Art_Mushroom.xlsx")

